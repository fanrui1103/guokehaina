# -*- coding: utf-8 -*-
"""益佳通标签表：装箱计算 + 生成 Excel。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

HEADERS = [
    "内箱标签序号",
    "供应商代码",
    "连接符1",
    "物料料号",
    "连接符2",
    "出货日期",
    "标签数据序号",
    "连接符3",
    "物料数量",
    "连接符4",
    "物料单位",
    "内箱标签二维码数据",
    "物料名称",
    "外箱标签连接符&",
    "外箱标签项",
    "外箱标签序号",
    "外箱物料数量",
    "外箱标签二维码数据",
    "卡板码连接符&",
    "卡板码标签项",
    "卡板码二维码数据",
]

COLUMN_WIDTHS = {
    "A": 12.2,
    "B": 10.88,
    "C": 6.62,
    "D": 18.26,
    "E": 6.61,
    "F": 10.88,
    "G": 11.46,
    "H": 7.2,
    "I": 8.88,
    "J": 6.62,
    "K": 8.88,
    "L": 40.88,
    "M": 12.0,
    "N": 15.73,
    "O": 46.13,
    "P": 12.79,
    "Q": 13.83,
    "R": 157.79,
    "S": 14.55,
    "T": 45.73,
    "U": 140.0,
}

INNER_PER_OUTER = 4
INNER_PER_PALLET_QR = 13
DASH = "-"
AMP = "&"
FONT = Font(name="宋体", size=11)
TEXT_FORMAT = numbers.FORMAT_TEXT
TEXT_COLUMNS = ("B", "C", "D", "E", "G", "H", "J", "K", "M")


def inner_quantities(total_qty: int, per_inner: int) -> list[int]:
    """按每箱件数拆成各内箱数量。除不尽时最后一箱装剩余件数。"""
    if total_qty <= 0:
        raise ValueError("总件数必须大于 0")
    if per_inner <= 0:
        raise ValueError("每内箱件数必须大于 0")
    n_full, remainder = divmod(total_qty, per_inner)
    qtys = [per_inner] * n_full
    if remainder:
        qtys.append(remainder)
    return qtys


def iter_groups(count: int, size: int):
    start = 0
    while start < count:
        end = min(start + size, count)
        yield start, end
        start = end


def packing_summary(total_qty: int, per_inner: int) -> dict[str, int]:
    qtys = inner_quantities(total_qty, per_inner)
    n_inner = len(qtys)
    n_outer = (n_inner + INNER_PER_OUTER - 1) // INNER_PER_OUTER
    n_pallet = (n_inner + INNER_PER_PALLET_QR - 1) // INNER_PER_PALLET_QR
    return {
        "inner": n_inner,
        "outer": n_outer,
        "pallet": n_pallet,
        "last_inner_qty": qtys[-1],
    }


def _excel_row(inner_index: int) -> int:
    return inner_index + 2


def _join_formula(col: str, start: int, end: int, op: str) -> str:
    """start/end 为内箱下标（含头不含尾）。op 为 + 或 &。"""
    parts = [f"{col}{_excel_row(i)}" for i in range(start, end)]
    return "=" + op.join(parts)


def _inner_concat_formula(row: int, extra_col: str) -> str:
    return (
        f"=B{row}&C{row}&D{row}&E{row}&F{row}&G{row}"
        f"&H{row}&I{row}&J{row}&K{row}&{extra_col}{row}"
    )


def build_workbook(
    *,
    supplier: str,
    part_no: str,
    ship_date: str | int,
    total_qty: int,
    per_inner: int,
    unit: str,
    name: str,
) -> Workbook:
    qtys = inner_quantities(total_qty, per_inner)
    n = len(qtys)
    serial_width = max(4, len(str(n)))

    last_of_outer = set()
    last_of_pallet = set()
    outer_groups = list(iter_groups(n, INNER_PER_OUTER))
    pallet_groups = list(iter_groups(n, INNER_PER_PALLET_QR))
    for start, end in outer_groups:
        last_of_outer.add(end - 1)
    for start, end in pallet_groups:
        last_of_pallet.add(end - 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "模板"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    try:
        date_value: str | int = int(str(ship_date).strip())
    except ValueError:
        date_value = str(ship_date).strip()

    for i, qty in enumerate(qtys):
        row = _excel_row(i)
        values = {
            "A": i + 1,
            "B": str(supplier).strip(),
            "C": DASH,
            "D": str(part_no).strip(),
            "E": DASH,
            "F": date_value,
            "G": f"{i + 1:0{serial_width}d}",
            "H": DASH,
            "I": qty,
            "J": DASH,
            "K": str(unit).strip(),
            "L": f"=B{row}&C{row}&D{row}&E{row}&F{row}&G{row}&H{row}&I{row}&J{row}&K{row}",
            "M": str(name).strip(),
            "N": None if i in last_of_outer else AMP,
            "O": _inner_concat_formula(row, "N"),
            "P": f"=A{row}",
            "S": None if i in last_of_pallet else AMP,
            "T": _inner_concat_formula(row, "S"),
        }
        for letter, value in values.items():
            cell = ws[f"{letter}{row}"]
            cell.value = value
            cell.font = FONT
            if letter in TEXT_COLUMNS:
                cell.number_format = TEXT_FORMAT

    for start, end in outer_groups:
        row = _excel_row(end - 1)
        q = ws[f"Q{row}"]
        r = ws[f"R{row}"]
        q.value = _join_formula("I", start, end, "+")
        r.value = _join_formula("O", start, end, "&")
        q.font = FONT
        r.font = FONT

    for start, end in pallet_groups:
        row = _excel_row(end - 1)
        u = ws[f"U{row}"]
        u.value = _join_formula("T", start, end, "&")
        u.font = FONT

    return wb


def default_filename(part_no: str, ship_date: str | int, total_qty: int) -> str:
    safe = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in str(part_no).strip())
    return f"益佳通标签_{safe}_{ship_date}_{total_qty}件.xlsx"


def save_workbook(wb: Workbook, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
