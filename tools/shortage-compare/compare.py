# -*- coding: utf-8 -*-
"""把「欠料表」当底稿，对照库存和生产计划，标出不够库存 / 还需生产 / 还需排产多少。"""

from __future__ import annotations

import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CODE_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._-]{5,}")

CAT_STOCK_OK = "①库存就够交货"
CAT_PLANNED = "②库存不够，但已排产"
CAT_BOTH_SHORT = "③库存和已排产都不够"

NEW_COLS = [
    "交货分类",
    "成品库存",
    "不良品库存",
    "库存判断",
    "库存缺口",
    "已排产数量",
    "已入库数量",
    "要求交期",
    "制令单号",
    "还需生产",
    "还需排产数量",
    "标注说明",
]


def to_num(value) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("，", "").replace(" ", "").strip()
    if text in {"", "-", "nan", "None"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def norm_code(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "nat"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def extract_codes(value) -> list[str]:
    text = norm_code(value)
    if not text:
        return []
    found = CODE_RE.findall(text)
    extra = re.split(r"[/|;,，、\s]+", text)
    codes: list[str] = []
    seen: set[str] = set()
    for item in found + extra:
        item = norm_code(item)
        if len(item) < 6:
            continue
        if not re.fullmatch(r"[A-Za-z0-9._-]+", item):
            continue
        key = item.casefold()
        if key in seen:
            continue
        seen.add(key)
        codes.append(item)
    return codes


def detect_kind(df: pd.DataFrame) -> str:
    cols = {str(c).strip() for c in df.columns}
    if "本期结存" in cols and "仓库名称" in cols:
        return "inventory"
    if "制令单号" in cols or ("投产数量" in cols and "客户货号" in cols):
        return "plan"
    if "物料编码" in cols and ("总欠料" in cols or "抵扣后的最终欠料" in cols):
        return "shortage"
    return "unknown"


def _col(df: pd.DataFrame, *names: str) -> str | None:
    mapping = {str(c).strip(): c for c in df.columns}
    for name in names:
        if name in mapping:
            return mapping[name]
    return None


def _series(df: pd.DataFrame, *names: str, default=None):
    col = _col(df, *names)
    if col is None:
        if default is None:
            return pd.Series([None] * len(df), index=df.index)
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


def build_inventory_index(inv: pd.DataFrame) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    warehouse_col = _col(inv, "仓库名称")
    mat_col = _col(inv, "物料编号")
    cust_col = _col(inv, "参考客户货号")
    stock_col = _col(inv, "本期结存")
    if not warehouse_col or not mat_col or not stock_col:
        return index

    for _, row in inv.iterrows():
        warehouse = str(row[warehouse_col]).strip() if pd.notna(row[warehouse_col]) else ""
        if warehouse in {"", "总计", "nan"}:
            continue
        item = {
            "仓库": warehouse,
            "物料编号": norm_code(row[mat_col]),
            "本期结存": to_num(row[stock_col]),
        }
        codes = extract_codes(row[cust_col] if cust_col else None)
        codes.append(item["物料编号"])
        seen: set[str] = set()
        for code in codes:
            key = code.casefold()
            if key in seen:
                continue
            seen.add(key)
            index.setdefault(key, []).append(item)
    return index


def build_plan_index(plan: pd.DataFrame) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    if plan is None or plan.empty:
        return index
    cust_col = _col(plan, "客户货号")
    if not cust_col:
        return index

    wo_col = _col(plan, "制令单号")
    launch_col = _col(plan, "投产数量")
    inbound_col = _col(plan, "生产入库数")
    due_col = _col(plan, "要求交期")
    status_col = _col(plan, "完工状态")
    name_col = _col(plan, "中文名")
    pid_col = _col(plan, "产品编号")
    stock_col = _col(plan, "仓库结存")

    for _, row in plan.iterrows():
        item = {
            "制令单号": norm_code(row[wo_col]) if wo_col else "",
            "投产数量": to_num(row[launch_col]) if launch_col else 0.0,
            "生产入库数": to_num(row[inbound_col]) if inbound_col else 0.0,
            "要求交期": row[due_col] if due_col else None,
            "完工状态": "" if not status_col or pd.isna(row[status_col]) else str(row[status_col]).strip(),
            "中文名": "" if not name_col or pd.isna(row[name_col]) else str(row[name_col]).strip(),
            "产品编号": norm_code(row[pid_col]) if pid_col else "",
            "仓库结存": to_num(row[stock_col]) if stock_col else 0.0,
        }
        codes = extract_codes(row[cust_col])
        if item["产品编号"]:
            codes.append(item["产品编号"])
        seen: set[str] = set()
        for code in codes:
            key = code.casefold()
            if key in seen:
                continue
            seen.add(key)
            index.setdefault(key, []).append(item)
    return index


def _uniq_rows(rows: list[dict], keys: tuple[str, ...]) -> list[dict]:
    unique: list[dict] = []
    seen: set[tuple] = set()
    for row in rows:
        mark = tuple(row.get(k, "") for k in keys)
        if mark in seen:
            continue
        seen.add(mark)
        unique.append(row)
    return unique


def _fmt_date(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text.lower() in {"", "nan", "nat", "none"}:
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return text


def _fmt_qty(value: float) -> float | int:
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 3)


def stock_for(code: str, inv_index: dict[str, list[dict]]) -> tuple[float, float, bool]:
    rows = _uniq_rows(inv_index.get(code.casefold(), []), ("仓库", "物料编号"))
    finished = 0.0
    ng = 0.0
    for row in rows:
        if row["仓库"] == "成品仓":
            finished += row["本期结存"]
        elif row["仓库"] == "不良品仓":
            ng += row["本期结存"]
    return finished, ng, bool(rows)


def plan_for(code: str, plan_index: dict[str, list[dict]]) -> dict:
    rows = _uniq_rows(plan_index.get(code.casefold(), []), ("制令单号", "产品编号"))
    launched = sum(r["投产数量"] for r in rows)
    inbound = sum(r["生产入库数"] for r in rows)
    dues = [_fmt_date(r["要求交期"]) for r in rows if _fmt_date(r["要求交期"])]
    wos = [r["制令单号"] for r in rows if r["制令单号"]]
    return {
        "已排产数量": launched,
        "已入库数量": inbound,
        "已排未完工": max(0.0, launched - inbound),
        "要求交期": "、".join(dict.fromkeys(dues)),
        "制令单号": "、".join(dict.fromkeys(wos)),
        "有排产": bool(rows),
    }


def annotate_shortage(shortage: pd.DataFrame, inventory: pd.DataFrame, plan: pd.DataFrame | None) -> pd.DataFrame:
    result = shortage.copy()
    unnamed = [c for c in result.columns if str(c).startswith("Unnamed")]
    if unnamed:
        result = result.rename(columns={unnamed[0]: "补充说明"})

    inv_index = build_inventory_index(inventory)
    plan_index = build_plan_index(plan if plan is not None else pd.DataFrame())

    code_col = _col(result, "物料编码")
    if not code_col:
        raise ValueError("欠料表里找不到「物料编码」这一列，请确认上传的是欠料表。")
    demand_col = _col(result, "抵扣后的最终欠料", "总欠料")

    cat_list = []
    finished_list = []
    ng_list = []
    judge_list = []
    gap_list = []
    launched_list = []
    inbound_list = []
    due_list = []
    wo_list = []
    need_prod_list = []
    need_plan_list = []
    note_list = []

    for _, row in result.iterrows():
        code = norm_code(row[code_col])
        demand = abs(to_num(row[demand_col] if demand_col else 0))
        finished, ng, stock_hit = stock_for(code, inv_index)
        info = plan_for(code, plan_index)

        gap = max(0.0, demand - finished)
        need_schedule = max(0.0, gap - info["已排未完工"])

        if demand <= 0 or gap <= 0:
            category = CAT_STOCK_OK
            judge = "抵扣后无欠料" if demand <= 0 else "库存足够"
            need_prod = "否"
            if demand <= 0:
                note = f"{category}。抵扣后欠料为 0，暂无交货缺口。成品库存 {_fmt_qty(finished)}。"
            else:
                note = f"{category}。成品库存 {_fmt_qty(finished)} ≥ 需求 {_fmt_qty(demand)}，现货可交。"
        elif need_schedule <= 0:
            category = CAT_PLANNED
            judge = "不够库存"
            need_prod = "是"
            note = (
                f"{category}。成品库存 {_fmt_qty(finished)}，需求 {_fmt_qty(demand)}，缺口 {_fmt_qty(gap)}。"
                f"已排产未入库 {_fmt_qty(info['已排未完工'])}，能盖住缺口，不用再开新单，但还要等生产。"
            )
        else:
            category = CAT_BOTH_SHORT
            judge = "不够库存"
            need_prod = "是"
            if info["有排产"]:
                note = (
                    f"{category}。成品库存 {_fmt_qty(finished)} + 已排未入库 {_fmt_qty(info['已排未完工'])}"
                    f" 仍小于需求 {_fmt_qty(demand)}，还需排产 {_fmt_qty(need_schedule)}。"
                )
            elif not stock_hit:
                note = (
                    f"{category}。库存表和生产计划都未匹配到该物料。"
                    f"缺口 {_fmt_qty(gap)}，还需排产 {_fmt_qty(need_schedule)}。"
                )
            else:
                note = (
                    f"{category}。成品库存 {_fmt_qty(finished)}，需求 {_fmt_qty(demand)}，缺口 {_fmt_qty(gap)}。"
                    f"生产计划中没有该物料，还需排产 {_fmt_qty(need_schedule)}。"
                )

        cat_list.append(category)
        finished_list.append(_fmt_qty(finished))
        ng_list.append(_fmt_qty(ng))
        judge_list.append(judge)
        gap_list.append(_fmt_qty(gap))
        launched_list.append(_fmt_qty(info["已排产数量"]))
        inbound_list.append(_fmt_qty(info["已入库数量"]))
        due_list.append(info["要求交期"])
        wo_list.append(info["制令单号"])
        need_prod_list.append(need_prod)
        need_plan_list.append(_fmt_qty(need_schedule))
        note_list.append(note)

    result["交货分类"] = cat_list
    result["成品库存"] = finished_list
    result["不良品库存"] = ng_list
    result["库存判断"] = judge_list
    result["库存缺口"] = gap_list
    result["已排产数量"] = launched_list
    result["已入库数量"] = inbound_list
    result["要求交期"] = due_list
    result["制令单号"] = wo_list
    result["还需生产"] = need_prod_list
    result["还需排产数量"] = need_plan_list
    result["标注说明"] = note_list

    # 交货分类放在原表列的后面、明细列的最前面
    original_cols = [c for c in result.columns if c not in NEW_COLS]
    result = result[original_cols + [c for c in NEW_COLS if c in result.columns]]
    return result


CAT_FILLS = {
    CAT_STOCK_OK: PatternFill("solid", fgColor="D8F3DC"),
    CAT_PLANNED: PatternFill("solid", fgColor="FFE8CC"),
    CAT_BOTH_SHORT: PatternFill("solid", fgColor="F8D7DA"),
}


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    headers = [cell.value for cell in ws[1]]
    cat_idx = headers.index("交货分类") + 1 if "交货分类" in headers else None
    note_idx = headers.index("标注说明") + 1 if "标注说明" in headers else None

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        category = str(ws.cell(row[0].row, cat_idx).value) if cat_idx else ""
        row_fill = CAT_FILLS.get(category, PatternFill("solid", fgColor="FFFFFF"))
        for cell in row:
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="微软雅黑", size=10)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28

    widths = {}
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 10
        for cell in col[:80]:
            longest = max(longest, min(len(str(cell.value or "")), 42))
        widths[letter] = longest + 2
    if note_idx:
        widths[get_column_letter(note_idx)] = 48
    if cat_idx:
        widths[get_column_letter(cat_idx)] = 24
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    sheets = [
        ("全部", df),
        ("①库存就够交货", df[df["交货分类"] == CAT_STOCK_OK] if "交货分类" in df.columns else df.iloc[0:0]),
        ("②库存不够但已排产", df[df["交货分类"] == CAT_PLANNED] if "交货分类" in df.columns else df.iloc[0:0]),
        ("③库存和已排产都不够", df[df["交货分类"] == CAT_BOTH_SHORT] if "交货分类" in df.columns else df.iloc[0:0]),
    ]
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, part in sheets:
            part.to_excel(writer, index=False, sheet_name=name)

    buf.seek(0)
    wb = load_workbook(buf)
    for ws in wb.worksheets:
        _style_sheet(ws)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def compare_files(shortage_file, inventory_file, plan_file=None) -> tuple[pd.DataFrame, dict]:
    shortage = pd.read_excel(shortage_file)
    inventory = pd.read_excel(inventory_file)
    plan = pd.read_excel(plan_file) if plan_file else None

    kinds = {
        "欠料表": detect_kind(shortage),
        "库存表": detect_kind(inventory),
        "生产计划": detect_kind(plan) if plan is not None else "未上传",
    }
    if kinds["欠料表"] != "shortage":
        raise ValueError("第一份表看起来不是欠料表（需要有「物料编码」和「总欠料」这类列）。")
    if kinds["库存表"] != "inventory":
        raise ValueError("第二份表看起来不是库存汇总表（需要有「仓库名称」和「本期结存」）。")
    if plan is not None and kinds["生产计划"] != "plan":
        raise ValueError("第三份表看起来不是主生产计划（需要有「制令单号」或「投产数量」）。")

    result = annotate_shortage(shortage, inventory, plan)
    cats = result["交货分类"] if "交货分类" in result.columns else pd.Series(dtype=str)
    summary = {
        "欠料行数": int(len(result)),
        CAT_STOCK_OK: int((cats == CAT_STOCK_OK).sum()),
        CAT_PLANNED: int((cats == CAT_PLANNED).sum()),
        CAT_BOTH_SHORT: int((cats == CAT_BOTH_SHORT).sum()),
        "文件识别": kinds,
    }
    return result, summary
