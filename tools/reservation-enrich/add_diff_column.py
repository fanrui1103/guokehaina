"""在智能排产合并结果上，于「订单数量」右侧插入「投产减入库」列。"""

from __future__ import annotations

import re
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

import pandas as pd
from openpyxl import load_workbook

NEW_COL = "投产减入库"
ORDER_COL = "订单数量"
SKU_COL = "供应料号"
MPS_SKU_COL = "客户货号"
MPS_START_COL = "投产数量"
MPS_IN_COL = "生产入库数"
MPS_STATUS_COL = "完工状态"
UNFINISHED_STATUS = "未完工"


def _to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _norm_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def read_merged_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        # 微信/系统导出常见 gbk；也兼容 utf-8
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                return pd.read_csv(path, dtype=str, encoding=enc)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(path, dtype=str)
    if suffix in {".xlsx", ".xlsm"}:
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    raise ValueError(f"合并表暂不支持该格式：{suffix}（请用 .csv 或 .xlsx）")


def build_diff_map_by_sku(mps_path: Path) -> dict[str, float]:
    """从主生产计划按客户货号汇总：Σ(投产数量 − 生产入库数)。

    只统计「未完工」制令单；已完工不计入。
    只按「客户货号」汇总，避免同一产品编号下不同料号互相串数。
    """
    wb = load_workbook(mps_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = ["" if h is None else str(h).strip() for h in header_row]
    try:
        i_sku = headers.index(MPS_SKU_COL)
        i_start = headers.index(MPS_START_COL)
        i_in = headers.index(MPS_IN_COL)
        i_status = headers.index(MPS_STATUS_COL)
    except ValueError as exc:
        wb.close()
        raise ValueError(
            "主生产计划缺少必要列，需要包含："
            f"{MPS_SKU_COL}、{MPS_START_COL}、{MPS_IN_COL}、{MPS_STATUS_COL}"
        ) from exc

    by_sku: dict[str, float] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        # openpyxl 会按最大列对齐；不足时补空
        cells = list(row)
        while len(cells) < len(headers):
            cells.append(None)

        status = _norm_key(cells[i_status])
        if status != UNFINISHED_STATUS:
            continue

        sku = _norm_key(cells[i_sku])
        if not sku:
            continue
        diff = _to_number(cells[i_start]) - _to_number(cells[i_in])
        by_sku[sku] = by_sku.get(sku, 0.0) + diff

    wb.close()
    return by_sku


def lookup_diff(row: pd.Series, by_sku: dict[str, float]) -> Any:
    """合并表「供应料号」对应主生产计划「客户货号」。"""
    sku = _norm_key(row.get(SKU_COL, ""))
    if sku and sku in by_sku:
        return by_sku[sku]
    return pd.NA


def insert_diff_column(merged: pd.DataFrame, mps_path: Path) -> pd.DataFrame:
    if ORDER_COL not in merged.columns:
        raise ValueError(f"合并表里找不到「{ORDER_COL}」列，请确认选的是智能排产合并结果")
    if SKU_COL not in merged.columns:
        raise ValueError(f"合并表里找不到「{SKU_COL}」列，请确认选的是智能排产合并结果")

    by_sku = build_diff_map_by_sku(mps_path)
    result = merged.copy()

    if NEW_COL in result.columns:
        result = result.drop(columns=[NEW_COL])

    values = result.apply(lambda r: lookup_diff(r, by_sku), axis=1)
    insert_at = list(result.columns).index(ORDER_COL) + 1
    result.insert(insert_at, NEW_COL, values)
    return result


def default_output_path(merged_path: Path) -> Path:
    stem = re.sub(r"-带投产减入库$", "", merged_path.stem)
    return merged_path.with_name(f"{stem}-带投产减入库.xlsx")


def save_result(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # 数值列尽量写成数字，方便 Excel 筛选
    out = df.copy()
    if NEW_COL in out.columns:
        out[NEW_COL] = pd.to_numeric(out[NEW_COL], errors="coerce")
    out.to_excel(output_path, index=False, engine="openpyxl")


def run_job(merged_path: Path, mps_path: Path, output_path: Path | None = None) -> Path:
    merged = read_merged_table(merged_path)
    result = insert_diff_column(merged, mps_path)
    out = output_path or default_output_path(merged_path)
    save_result(result, out)
    return out


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("排产结果 · 增加投产减入库列")
        self.geometry("640x260")
        self.resizable(False, False)

        self.merged_var = tk.StringVar()
        self.mps_var = tk.StringVar()

        pad = {"padx": 12, "pady": 8}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, **pad)

        ttk.Label(frm, text="1. 合并后的表（智能排产结果，csv/xlsx）").grid(
            row=0, column=0, columnspan=3, sticky="w"
        )
        ttk.Entry(frm, textvariable=self.merged_var, width=62).grid(row=1, column=0, columnspan=2, sticky="we")
        ttk.Button(frm, text="选择…", command=self.pick_merged).grid(row=1, column=2, padx=6)

        ttk.Label(frm, text="2. 主生产计划（xlsx）").grid(row=2, column=0, columnspan=3, sticky="w", pady=(12, 0))
        ttk.Entry(frm, textvariable=self.mps_var, width=62).grid(row=3, column=0, columnspan=2, sticky="we")
        ttk.Button(frm, text="选择…", command=self.pick_mps).grid(row=3, column=2, padx=6)

        ttk.Button(frm, text="生成", command=self.on_generate).grid(row=4, column=0, sticky="w", pady=18)
        ttk.Label(
            frm,
            text="新列「投产减入库」= 未完工制令的(投产−入库)；按供应料号=客户货号；对不上留空",
            foreground="#444",
        ).grid(row=5, column=0, columnspan=3, sticky="w")

        frm.columnconfigure(0, weight=1)

    def pick_merged(self) -> None:
        path = filedialog.askopenfilename(
            title="选择合并后的表",
            filetypes=[
                ("表格", "*.csv;*.xlsx;*.xlsm"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx;*.xlsm"),
                ("全部", "*.*"),
            ],
        )
        if path:
            self.merged_var.set(path)

    def pick_mps(self) -> None:
        path = filedialog.askopenfilename(
            title="选择主生产计划",
            filetypes=[("Excel", "*.xlsx;*.xlsm"), ("全部", "*.*")],
        )
        if path:
            self.mps_var.set(path)

    def on_generate(self) -> None:
        merged = self.merged_var.get().strip()
        mps = self.mps_var.get().strip()
        if not merged or not mps:
            messagebox.showwarning("提示", "请先选择合并表和主生产计划两个文件")
            return
        merged_path = Path(merged)
        mps_path = Path(mps)
        if not merged_path.is_file():
            messagebox.showerror("错误", f"找不到合并表：\n{merged_path}")
            return
        if not mps_path.is_file():
            messagebox.showerror("错误", f"找不到主生产计划：\n{mps_path}")
            return

        try:
            out = run_job(merged_path, mps_path)
        except Exception as exc:  # noqa: BLE001 - 给使用者看清楚原因
            messagebox.showerror("生成失败", str(exc))
            return

        open_folder = messagebox.askyesno(
            "完成",
            f"已生成（未覆盖原文件）：\n{out}\n\n要打开所在文件夹吗？",
        )
        if open_folder:
            try:
                import os

                os.startfile(out.parent)  # type: ignore[attr-defined]
            except OSError:
                pass


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
