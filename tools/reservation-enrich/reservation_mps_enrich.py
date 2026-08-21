"""预约与库存报表 + 主生产计划：原表保留，插入订单数量/仓库结存/投产减入库/提示。

用法：选「预约与库存报表」+「主生产计划」→ 开始处理 → 导出 xlsx。
新增列插在「5周订货量」右侧，顺序：订单数量 → 仓库结存 → 投产减入库 → 提示。
导出只保留「需求」行；相同供应料号排在一起。
料号先后顺序跟客户系统导出的预约报表一致（越靠前越急），表格每天变就跟着变。

提示按料号汇总：
  5周合计 = 各地点「5周订货量」相加
  已预约合计 = 各地点「已预约交货量」相加
  仓库结存、投产减入库：同一料号只算一次
  覆盖量 = 已预约合计 + 仓库结存 + 投产减入库（后两项同料号只算一次）
  欠量 = 5周合计 − 覆盖量
    欠量 > 0 →「欠xxx」(黄)；5周合计 > 覆盖×1.2 →「超20%，欠xxx」(红)
  多出 = 覆盖量 − 5周合计
    5周合计 > 0 且 多出 > 5周合计×1.2 →「库存预警，多xxx」(蓝)，需跟进需求进度
  提示文字只写在该料号的第一行；该料号各地点行都涂同色，方便按颜色筛选不漏地址。

排版：正文默认水平+垂直居中、全表细边框、列宽按内容拉开（避免 ######）。
"""

from __future__ import annotations

import traceback
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from smart_schedule import (
    _norm_key,
    _to_number,
    build_diff_map_by_sku,
    fuzzy_match,
    read_company,
    read_customer_raw,
)

APP_TITLE = "预约报表补数（提示版）"
NEW_COLS = ["订单数量", "仓库结存", "投产减入库", "提示"]

FILL_WARN = PatternFill(fill_type="solid", fgColor="FFF2CC")  # 黄：欠
FILL_ALERT = PatternFill(fill_type="solid", fgColor="F4CCCC")  # 红：超20%欠
FILL_STOCK = PatternFill(fill_type="solid", fgColor="DDEBF7")  # 蓝：库存预警（备太多）
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 品名规格：左对齐+换行；列够宽、行够高，整段规格都能看见
NAME_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
NAME_COL_WIDTH = 72  # 加宽，少截断
NAME_MAX_LINES = 8


def _name_row_height(text: Any) -> float:
    """按品名规格字数估算行高，保证换行后能看全。"""
    dlen = _display_len(text)
    if dlen <= 0:
        return 22
    # 偏保守：实际显示往往比列宽数字能装的更少（字体/缩放）
    chars_per_line = max(int(NAME_COL_WIDTH * 0.7), 24)
    lines = max(1, (dlen + chars_per_line - 1) // chars_per_line)
    lines = min(lines, NAME_MAX_LINES)
    # 超过一列宽的内容至少给两行高度
    if dlen > chars_per_line:
        lines = max(lines, 2)
    return float(max(24, min(lines * 17 + 6, 150)))


def _cell_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _fmt_qty(value: float) -> str:
    """数量显示：整数不带小数点。"""
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _excel_value(value: Any) -> Any:
    """写入 Excel 前的值：日期只要年月日，不要 0:00:00。"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    # pandas.Timestamp / numpy datetime
    if hasattr(value, "to_pydatetime"):
        try:
            value = value.to_pydatetime()
        except Exception:
            pass
    if isinstance(value, datetime):
        return f"{value.year}/{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.year}/{value.month}/{value.day}"
    return value


def find_columns(df: pd.DataFrame) -> dict[str, int]:
    """根据第2、3行表头定位关键列（0-based）。"""
    if df.shape[0] < 3:
        raise ValueError("预约与库存报表行数太少，请确认文件")

    found: dict[str, int] = {}
    for c in range(df.shape[1]):
        top = _cell_str(df.iloc[1, c])
        bot = _cell_str(df.iloc[2, c])
        if top == "序号":
            found["序号"] = c
        if top == "交货地点":
            found["交货地点"] = c
        if top == "供应料号":
            found["供应料号"] = c
        if bot == "交货量" and "已预约" in top:
            found["已预约交货量"] = c
        if bot == "订货量" and "5周" in top:
            found["5周订货量"] = c
        if bot == "类型":
            found["类型"] = c

    need = ["供应料号", "已预约交货量", "5周订货量", "类型"]
    missing = [k for k in need if k not in found]
    if missing:
        raise ValueError("预约报表缺少列：" + "、".join(missing) + "（请确认是标准「预约与库存报表」）")
    return found


def tip_for(week5: float, cover: float) -> tuple[str, str | None]:
    """返回 (提示文字, 颜色档位 alert/warn/stock/None)。

    欠料：5周 > 覆盖量
    库存预警：覆盖量比 5周 多出的部分 > 5周×1.2（备货过多，需跟进需求）
    """
    gap = week5 - cover
    if gap > 0:
        qty = _fmt_qty(gap)
        if week5 > cover * 1.2:
            return f"超20%，欠{qty}", "alert"
        return f"欠{qty}", "warn"

    # 备太多：多出 = 覆盖 − 5周；多出 > 5周×120%
    if week5 > 0:
        excess = cover - week5
        if excess > week5 * 1.2:
            return f"库存预警，多{_fmt_qty(excess)}", "stock"
    return "", None


def _display_len(value: Any) -> int:
    """估算单元格显示宽度（中文约按 2 个字符宽）。"""
    if value is None:
        return 0
    if isinstance(value, datetime):
        return 12
    text = str(value)
    width = 0
    for ch in text:
        width += 2 if ord(ch) > 127 else 1
    return width


def format_sheet(ws: Worksheet) -> None:
    """居中 + 边框 + 列宽，让导出表更好看、数字不被 ###### 挡住。"""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    # 找出「品名规格」列
    name_col = None
    for c in range(1, max_col + 1):
        if _cell_str(ws.cell(2, c).value) == "品名规格":
            name_col = c
            break

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if name_col and cell.column == name_col and cell.row >= 4:
                cell.alignment = NAME_ALIGN
            else:
                cell.alignment = CENTER
            cell.border = THIN_BORDER

    # 标题/表头行高；表头加粗浅蓝底
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    for c in range(1, max_col + 1):
        for r in (2, 3):
            cell = ws.cell(r, c)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL

    # 列宽：从表头/数据估算（跳过第1行大标题，避免把 A 列撑爆）
    for c in range(1, max_col + 1):
        if name_col and c == name_col:
            ws.column_dimensions[get_column_letter(c)].width = NAME_COL_WIDTH
            continue

        longest = 0
        sample_to = min(max_row, 100)
        for r in range(2, sample_to + 1):
            longest = max(longest, _display_len(ws.cell(r, c).value))

        width = min(max(longest + 2, 6), 18)
        sample = ws.cell(3, c).value
        # 日期表头：2026/8/3 这种短日期
        if isinstance(sample, str) and sample.count("/") >= 2 and sample[:4].isdigit():
            width = max(width, 11)
        elif isinstance(sample, (datetime, date)):
            width = max(width, 11)
        if _cell_str(ws.cell(2, c).value) == "提示":
            width = max(width, 20)
        ws.column_dimensions[get_column_letter(c)].width = width

    # 数据行高：按品名规格加高，把整段显示全
    for r in range(4, max_row + 1):
        if name_col:
            ws.row_dimensions[r].height = _name_row_height(ws.cell(r, name_col).value)
        else:
            ws.row_dimensions[r].height = 22

    # 顶栏标题：跨全部列合并居中（类似点「合并居中」）
    title = ws.cell(1, 1).value
    if title not in (None, ""):
        for c in range(2, max_col + 1):
            ws.cell(1, c).value = None
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(1, 1)
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14)
        title_cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 28

    # 冻结表头，翻页时列名还在
    ws.freeze_panes = "A4"


def enrich_and_export(customer_path: Path, company_path: Path, export_path: Path) -> Path:
    raw = read_customer_raw(customer_path)
    cols = find_columns(raw)
    insert_at = cols["5周订货量"] + 1  # 插在 5周订货量 右侧
    n_new = len(NEW_COLS)

    company = read_company(company_path)
    needed = ["制令单号", "客户货号", "完工状态", "订单数量", "仓库结存", "产品编号", "投产数量", "生产入库数"]
    missing = [c for c in needed if c not in company.columns]
    if missing:
        raise ValueError("主生产计划缺少列：" + "、".join(missing))

    company_sorted = company.sort_values("制令单号")
    diff_map = build_diff_map_by_sku(company)

    sku_c = cols["供应料号"]
    booked_c = cols["已预约交货量"]
    week5_c = cols["5周订货量"]
    type_c = cols["类型"]
    loc_c = cols.get("交货地点")
    seq_c = cols.get("序号")

    nrows, ncols = raw.shape

    # 收集需求行 + 匹配主生产计划
    records: list[dict[str, Any]] = []
    for r in range(4, nrows):
        if _cell_str(raw.iloc[r, type_c]) != "需求":
            continue
        sku = _norm_key(raw.iloc[r, sku_c])
        row_series = pd.Series({"供应料号": raw.iloc[r, sku_c]})
        matched = fuzzy_match(row_series, company_sorted)
        records.append(
            {
                "src_r": r,
                "sku": sku,
                "loc": _cell_str(raw.iloc[r, loc_c]) if loc_c is not None else "",
                "booked": _to_number(raw.iloc[r, booked_c]),
                "week5": _to_number(raw.iloc[r, week5_c]),
                "order_qty": _to_number(matched.get("订单数量", 0)),
                "warehouse": _to_number(matched.get("仓库结存", 0)),
                "prod_diff": float(diff_map[sku]) if sku and sku in diff_map else 0.0,
            }
        )

    # 按料号汇总提示：已预约/5周各地相加；仓库结存、投产减入库只取一次
    sum_booked: dict[str, float] = defaultdict(float)
    sum_week5: dict[str, float] = defaultdict(float)
    once_wh: dict[str, float] = {}
    once_diff: dict[str, float] = {}
    for rec in records:
        sku = rec["sku"]
        sum_booked[sku] += rec["booked"]
        sum_week5[sku] += rec["week5"]
        if sku not in once_wh:
            once_wh[sku] = rec["warehouse"]
            once_diff[sku] = rec["prod_diff"]

    sku_tips: dict[str, tuple[str, str | None]] = {}
    for sku in sum_week5:
        cover = sum_booked[sku] + once_wh[sku] + once_diff[sku]
        sku_tips[sku] = tip_for(sum_week5[sku], cover)

    # 料号组顺序：跟客户原表首次出现顺序一致（越靠前越急）；组内同料号聚在一起，再按交货地点
    sku_rank: dict[str, int] = {}
    for rec in records:
        if rec["sku"] not in sku_rank:
            sku_rank[rec["sku"]] = len(sku_rank)
    records.sort(key=lambda x: (sku_rank.get(x["sku"], 10**9), x["loc"], x["src_r"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "预约与库存报表"

    def write_row(src_r: int, out_r: int, fill_new_header: bool = False) -> None:
        out_c = 0
        for c in range(ncols):
            if c == insert_at:
                for i, name in enumerate(NEW_COLS):
                    cell = ws.cell(row=out_r, column=out_c + 1)
                    if fill_new_header and src_r == 1:
                        cell.value = name
                        cell.fill = HEADER_FILL
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    else:
                        cell.value = None
                    out_c += 1
            val = _excel_value(raw.iloc[src_r, c])
            ws.cell(row=out_r, column=out_c + 1, value=val)
            out_c += 1

    write_row(0, 1, fill_new_header=False)
    write_row(1, 2, fill_new_header=True)
    write_row(2, 3, fill_new_header=False)

    new_order_col = insert_at + 1  # 1-based
    new_wh_col = insert_at + 2
    new_diff_col = insert_at + 3
    new_tip_col = insert_at + 4
    last_col = ncols + n_new

    def out_col_of(src_c: int) -> int:
        """原表 0-based 列 → 导出 1-based 列（已插入新列）。"""
        if src_c >= insert_at:
            return src_c + n_new + 1
        return src_c + 1

    tipped_skus: set[str] = set()
    out_r = 4
    seq = 1
    for rec in records:
        write_row(rec["src_r"], out_r, fill_new_header=False)

        if seq_c is not None:
            ws.cell(row=out_r, column=out_col_of(seq_c), value=seq)

        ws.cell(row=out_r, column=new_order_col, value=rec["order_qty"])
        ws.cell(row=out_r, column=new_wh_col, value=rec["warehouse"])
        ws.cell(row=out_r, column=new_diff_col, value=rec["prod_diff"])

        tip_text, level = sku_tips.get(rec["sku"], ("", None))
        # 文字仍只写第一行；颜色涂满该料号所有交货地点行，按颜色筛选不会漏
        is_first = rec["sku"] not in tipped_skus
        if is_first:
            tipped_skus.add(rec["sku"])
        tip = tip_text if is_first else ""

        tip_cell = ws.cell(row=out_r, column=new_tip_col, value=tip or None)
        if tip:
            tip_cell.font = Font(bold=True)

        if level == "alert":
            fill = FILL_ALERT
        elif level == "warn":
            fill = FILL_WARN
        elif level == "stock":
            fill = FILL_STOCK
        else:
            fill = None
        if fill is not None:
            for col_i in range(1, last_col + 1):
                ws.cell(row=out_r, column=col_i).fill = fill

        out_r += 1
        seq += 1

    format_sheet(ws)

    export_path = export_path if export_path.suffix.lower() == ".xlsx" else export_path.with_suffix(".xlsx")
    export_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(export_path)
    return export_path


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("780x420")
        self.resizable(False, False)

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.export_path = tk.StringVar(value="选择导出路径")

        pad = {"padx": 12, "pady": 6}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, **pad)

        ttk.Label(frm, text="客户文件（预约与库存报表 .xls/.xlsx）：").grid(
            row=0, column=0, columnspan=3, sticky="w"
        )
        ttk.Entry(frm, textvariable=self.file1_path, width=72).grid(row=1, column=0, columnspan=2, sticky="we")
        ttk.Button(frm, text="浏览", command=lambda: self.select_file(self.file1_path)).grid(
            row=1, column=2, padx=6
        )

        ttk.Label(frm, text="公司文件（主生产计划 .xlsx）：").grid(
            row=2, column=0, columnspan=3, sticky="w", pady=(10, 0)
        )
        ttk.Entry(frm, textvariable=self.file2_path, width=72).grid(row=3, column=0, columnspan=2, sticky="we")
        ttk.Button(frm, text="浏览", command=lambda: self.select_file(self.file2_path)).grid(
            row=3, column=2, padx=6
        )

        ttk.Label(frm, text="导出路径：").grid(row=4, column=0, columnspan=3, sticky="w", pady=(10, 0))
        ttk.Entry(frm, textvariable=self.export_path, width=72).grid(row=5, column=0, columnspan=2, sticky="we")
        ttk.Button(frm, text="选择路径", command=self.select_export_path).grid(row=5, column=2, padx=6)

        ttk.Button(frm, text="开始处理", command=self.process_files).grid(row=6, column=0, sticky="w", pady=16)
        ttk.Label(
            frm,
            text="说明：顺序跟客户预约报表一致（越前越急）；同料号仍排一起并同色标记。"
            "导出居中、边框、日期只要年月日。",
            foreground="#444",
            wraplength=720,
        ).grid(row=7, column=0, columnspan=3, sticky="w")

        self.status = ttk.Label(frm, text="准备就绪", relief=tk.SUNKEN, anchor="w")
        self.status.grid(row=8, column=0, columnspan=3, sticky="we", pady=(18, 0))
        frm.columnconfigure(0, weight=1)

    def select_file(self, var: tk.StringVar) -> None:
        path = filedialog.askopenfilename(
            title="选择文件",
            filetypes=[("Excel", "*.xls;*.xlsx;*.xlsm"), ("全部", "*.*")],
        )
        if path:
            var.set(path)

    def select_export_path(self) -> None:
        default_name = datetime.now().strftime("%Y-%m-%d") + "-预约报表补数.xlsx"
        path = filedialog.asksaveasfilename(
            title="选择导出路径",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("全部", "*.*")],
        )
        if path:
            self.export_path.set(path)

    def process_files(self) -> None:
        f1 = self.file1_path.get().strip()
        f2 = self.file2_path.get().strip()
        export = self.export_path.get().strip()
        if not f1 or not f2:
            messagebox.showwarning("提示", "请先选择客户文件和公司文件")
            return
        if not export or export == "选择导出路径":
            messagebox.showwarning("提示", "请先选择导出路径")
            return
        try:
            self.status.config(text="处理中，请稍候…")
            self.update_idletasks()
            out = enrich_and_export(Path(f1), Path(f2), Path(export))
            self.status.config(text="处理完成")
            messagebox.showinfo("成功", f"文件已成功导出到：\n{out}")
        except Exception as exc:  # noqa: BLE001
            traceback.print_exc()
            self.status.config(text="错误发生")
            messagebox.showerror("错误", f"处理过程中发生错误：\n{exc}")


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
