# -*- coding: utf-8 -*-
"""怡富万物料标示单批量生成：从你提供的模板读取内容，只改序列号和二维码。"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import qrcode
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

FONT_NAME = "楷体"
THIN = Side(style="thin", color="000000")

LABEL_START_COLS = (2, 7, 12)  # B / G / L
BLOCK_STRIDE = 13
LABELS_PER_ROW = 3
ROWS_PER_PAGE = 4
LABELS_PER_SHEET = LABELS_PER_ROW * ROWS_PER_PAGE

COL_WIDTHS = {
    "A": 0.5,
    "B": 8.875,
    "C": 17.5,
    "D": 8.75,
    "E": 7.875,
    "F": 1.125,
    "G": 8.875,
    "H": 17.5,
    "I": 8.75,
    "J": 7.875,
    "K": 1.625,
    "L": 8.875,
    "M": 17.5,
    "N": 8.75,
    "O": 7.875,
}

ROW_HEIGHTS = {
    0: 23.25,
    1: 16.5,
    2: 20.25,
    3: 16.5,
    4: 24.95,
    5: 24.95,
    6: 24.95,
    7: 20.0,
    8: 20.0,
    9: 20.0,
    10: 20.0,
    11: 21.75,
    12: 7.5,
}

QR_FROM_COLOFF = 171450
QR_FROM_ROWOFF = 19050
QR_TO_COLOFF = 419735
QR_TO_ROWOFF = 257810

# 左列字段相对标题行的行偏移（旧料号保留空行，不要求填写）
FIELD_OFFSETS = {
    "vendor": 1,
    "material_no": 2,
    "old_material_no": 3,
    "qty": 4,
    "product_name": 5,
    "batch_no": 6,
    "serial": 7,
    "delivery_date": 8,
    "production_date": 9,
    "expiry_date": 10,
    "order_no": 11,
}


@dataclass
class TemplateInfo:
    """从模板第一张标签读出来的内容。"""

    source_path: str
    sheet_name: str
    title: str = "怡富萬電業（惠州）有限公司——物料標示單"
    field_labels: dict[int, str] = field(default_factory=dict)
    vendor: str = ""
    material_no: str = ""
    qty_text: str = ""
    product_name: str = ""
    batch_no: str = ""
    delivery_date: str = ""
    production_date: str = ""
    expiry_date: str = ""
    order_no: str = ""
    rohs: str = "RoHS"
    iqc: str = "IQC"
    # 二维码：除最后一段序列号外，整段照抄模板
    qr_prefix_parts: list[str] = field(default_factory=list)
    sample_qr_payload: str = ""
    sample_display_serial: str = ""
    sample_qr_serial: str = ""
    display_serial_prefix: str = "OO"  # OO 或空
    display_serial_width: int = 2
    qr_serial_width: int = 4


@dataclass
class GenerateRequest:
    template: TemplateInfo
    serials: list[int]
    sheet_name: str = ""


def parse_serial_spec(text: str) -> list[int]:
    """把用户填写的序号规则解析成整数列表。

    支持示例：
    - 1-24
    - 1-24, 1-24, 1-24
    - 1-10, 15, 20-25
    - 多行书写也可以
    分隔符：逗号、中文逗号、分号、空白换行均可。
    """
    raw = (text or "").strip()
    if not raw:
        raise ValueError("请填写要生成的序列号，例如：1-24  或  1-24,1-24,1-12")

    # 统一分隔
    normalized = raw.replace("，", ",").replace("；", ",").replace(";", ",")
    normalized = normalized.replace("\n", ",").replace("\r", ",").replace("、", ",")
    parts = [p.strip() for p in normalized.split(",") if p.strip()]
    if not parts:
        raise ValueError("没有解析到有效序号，请检查填写")

    serials: list[int] = []
    for part in parts:
        token = part.replace("～", "-").replace("~", "-").replace("—", "-").replace("–", "-")
        if re.fullmatch(r"\d+", token):
            n = int(token)
            if n < 1:
                raise ValueError(f"序号必须 ≥ 1：{part}")
            serials.append(n)
            continue
        m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", token)
        if not m:
            raise ValueError(f"看不懂这段序号「{part}」。请写成 1-24 或单个数字 15")
        a, b = int(m.group(1)), int(m.group(2))
        if a < 1 or b < 1:
            raise ValueError(f"序号必须 ≥ 1：{part}")
        if a > b:
            raise ValueError(f"区间写反了「{part}」，应从小到大，例如 1-24")
        if b - a > 5000:
            raise ValueError(f"区间「{part}」太大（超过 5000），请拆开填写")
        serials.extend(range(a, b + 1))

    if not serials:
        raise ValueError("解析结果为空，请重新填写")
    return serials


def summarize_serials(serials: list[int], limit_chunks: int = 20) -> str:
    """把序号列表压成简短说明，方便预览。"""
    if not serials:
        return "（无）"
    chunks: list[str] = []
    start = prev = serials[0]
    for n in serials[1:]:
        if n == prev + 1:
            prev = n
            continue
        chunks.append(str(start) if start == prev else f"{start}-{prev}")
        start = prev = n
    chunks.append(str(start) if start == prev else f"{start}-{prev}")
    if len(chunks) > limit_chunks:
        shown = ", ".join(chunks[:limit_chunks])
        return f"共 {len(serials)} 个：{shown}, ..."
    return f"共 {len(serials)} 个：{', '.join(chunks)}"


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _digits(text: str) -> str:
    return re.sub(r"\D", "", str(text or ""))


def _as_int_or_str(text: str):
    s = str(text).strip()
    if re.fullmatch(r"\d+", s):
        return int(s)
    return s if s else None


def _parse_serial_number(text: str) -> int | None:
    s = _cell_str(text)
    if not s:
        return None
    m = re.search(r"(\d+)$", s)
    if not m:
        return None
    return int(m.group(1))


def display_serial(info: TemplateInfo, n: int) -> str:
    body = f"{n:0{info.display_serial_width}d}"
    return f"{info.display_serial_prefix}{body}"


def qr_serial(info: TemplateInfo, n: int) -> str:
    return f"{n:0{info.qr_serial_width}d}"


def build_qr_payload(info: TemplateInfo, serial_no: int) -> str:
    if info.qr_prefix_parts:
        return ";".join([*_cell_parts(info.qr_prefix_parts), qr_serial(info, serial_no)])
    # 兜底：没有扫到模板二维码时，用格子拼
    qty_num = _digits(info.qty_text) or info.qty_text
    vendor_code = info.vendor.split("-", 1)[0].strip() if "-" in info.vendor else info.vendor
    return ";".join(
        [
            vendor_code,
            info.material_no,
            qty_num,
            _digits(info.production_date),
            _digits(info.expiry_date),
            info.order_no,
            info.rohs or "RoHS",
            _digits(info.batch_no) or info.batch_no,
            qr_serial(info, serial_no),
        ]
    )


def _cell_parts(parts: list[str]) -> list[str]:
    return [str(p) for p in parts]


def make_qr_png_bytes(payload: str, box_size: int = 6, border: int = 2) -> bytes:
    from PIL import Image as PILImage

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=box_size,
        border=border,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = img.resize((300, 300), resample=PILImage.NEAREST)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _decode_png_bytes(data: bytes) -> str:
    """尽量读出二维码文字。优先 pyzbar，其次 OpenCV。"""
    try:
        from PIL import Image
        from pyzbar.pyzbar import decode

        vals = decode(Image.open(io.BytesIO(data)))
        if vals:
            return vals[0].data.decode("utf-8", errors="replace")
    except Exception:
        pass
    try:
        import cv2
        import numpy as np

        arr = np.frombuffer(data, dtype=np.uint8)
        im = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if im is None:
            return ""
        val, _, _ = cv2.QRCodeDetector().detectAndDecode(im)
        if val:
            return val
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        big = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_NEAREST)
        val, _, _ = cv2.QRCodeDetector().detectAndDecode(big)
        return val or ""
    except Exception:
        return ""


def list_label_sheets(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = []
        for name in wb.sheetnames:
            ws = wb[name]
            # 粗略判断：前几行出现「序列号」或标题含「物料」
            hit = False
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=True):
                for v in row:
                    s = _cell_str(v)
                    if "序列号" in s or "物料標示" in s or "物料标示" in s:
                        hit = True
                        break
                if hit:
                    break
            if hit:
                names.append(name)
        return names or list(wb.sheetnames)
    finally:
        wb.close()


def _read_first_qr_payload(ws) -> str:
    # 模板里常有大量残留图，只试前若干张，优先锚在第一块标签附近的
    candidates = []
    for img in ws._images:
        try:
            row = img.anchor._from.row
            col = img.anchor._from.col
        except Exception:
            row, col = 999, 999
        candidates.append((row, col, img))
    candidates.sort(key=lambda x: (x[0], x[1]))

    for row, col, img in candidates[:40]:
        try:
            raw = img._data()
        except Exception:
            continue
        text = _decode_png_bytes(raw)
        if text and ";" in text:
            return text
    return ""


def load_template(path: str | Path, sheet_name: str | None = None) -> TemplateInfo:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"找不到模板：{path}")

    wb = load_workbook(path, data_only=False)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在：{sheet_name}")
            ws = wb[sheet_name]
        else:
            sheets = list_label_sheets(path)
            ws = wb[sheets[0]]
            sheet_name = ws.title

        # 第一张标签固定读 B2 区域
        sr, sc = 2, 2
        c_val = sc + 1
        c_right = sc + 2

        title = _cell_str(ws.cell(sr, sc).value) or "怡富萬電業（惠州）有限公司——物料標示單"
        field_labels = {}
        for off in range(1, 12):
            field_labels[off] = _cell_str(ws.cell(sr + off, sc).value)

        vendor = _cell_str(ws.cell(sr + FIELD_OFFSETS["vendor"], c_val).value)
        material_no = _cell_str(ws.cell(sr + FIELD_OFFSETS["material_no"], c_val).value)
        qty_text = _cell_str(ws.cell(sr + FIELD_OFFSETS["qty"], c_val).value)
        product_name = _cell_str(ws.cell(sr + FIELD_OFFSETS["product_name"], c_val).value)
        batch_no = _cell_str(ws.cell(sr + FIELD_OFFSETS["batch_no"], c_val).value)
        sample_display = _cell_str(ws.cell(sr + FIELD_OFFSETS["serial"], c_val).value)
        delivery_date = _cell_str(ws.cell(sr + FIELD_OFFSETS["delivery_date"], c_val).value)
        production_date = _cell_str(ws.cell(sr + FIELD_OFFSETS["production_date"], c_val).value)
        expiry_date = _cell_str(ws.cell(sr + FIELD_OFFSETS["expiry_date"], c_val).value)
        order_no = _cell_str(ws.cell(sr + FIELD_OFFSETS["order_no"], c_val).value)
        rohs = _cell_str(ws.cell(sr + FIELD_OFFSETS["batch_no"], c_right).value) or "RoHS"
        iqc = _cell_str(ws.cell(sr + FIELD_OFFSETS["serial"], c_right).value) or "IQC"

        if not material_no and not product_name:
            raise ValueError(
                f"工作表「{sheet_name}」左上角不像物料标签，请换一个含标签的工作表。"
            )

        qr_payload = _read_first_qr_payload(ws)
        qr_prefix_parts: list[str] = []
        sample_qr_serial = ""
        if qr_payload:
            parts = qr_payload.split(";")
            if len(parts) >= 2:
                sample_qr_serial = parts[-1]
                qr_prefix_parts = parts[:-1]

        # 显示序列号格式：OO13 / 0013 / 13
        display_prefix = "OO"
        display_width = 2
        if sample_display:
            if sample_display.upper().startswith("OO") and _parse_serial_number(sample_display):
                display_prefix = sample_display[:2]
                display_width = max(2, len(_digits(sample_display)))
            elif re.fullmatch(r"\d+", sample_display):
                display_prefix = ""
                display_width = max(2, len(sample_display))
            else:
                # 其它前缀，尽量保留非数字前缀
                m = re.match(r"^(.*?)(\d+)$", sample_display)
                if m:
                    display_prefix = m.group(1)
                    display_width = max(2, len(m.group(2)))

        qr_width = max(4, len(sample_qr_serial)) if sample_qr_serial else 4

        return TemplateInfo(
            source_path=str(path),
            sheet_name=sheet_name,
            title=title,
            field_labels=field_labels,
            vendor=vendor,
            material_no=material_no,
            qty_text=qty_text,
            product_name=product_name,
            batch_no=batch_no,
            delivery_date=delivery_date,
            production_date=production_date,
            expiry_date=expiry_date,
            order_no=order_no,
            rohs=rohs,
            iqc=iqc,
            qr_prefix_parts=qr_prefix_parts,
            sample_qr_payload=qr_payload,
            sample_display_serial=sample_display,
            sample_qr_serial=sample_qr_serial,
            display_serial_prefix=display_prefix,
            display_serial_width=display_width,
            qr_serial_width=qr_width,
        )
    finally:
        wb.close()


def template_summary(info: TemplateInfo) -> str:
    lines = [
        f"模板文件：{info.source_path}",
        f"工作表：{info.sheet_name}",
        f"标题：{info.title}",
        f"厂商代码：{info.vendor}",
        f"料号：{info.material_no}",
        f"品名：{info.product_name}",
        f"数量：{info.qty_text}",
        f"批号：{info.batch_no}",
        f"交货日期：{info.delivery_date}",
        f"生产日期：{info.production_date}",
        f"有效日期：{info.expiry_date}",
        f"订单号：{info.order_no}",
        f"旧料号：不填写（留空）",
        f"模板样例序列号：{info.sample_display_serial or '（空）'}",
        f"模板样例二维码：{info.sample_qr_payload or '（未读到，将按格子拼）'}",
    ]
    return "\n".join(lines)


def _font(size: float, name: str = FONT_NAME) -> Font:
    return Font(name=name, size=size)


def _align(h=None, v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_border_range(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _merge(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    if r1 == r2 and c1 == c2:
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _default_field_labels() -> dict[int, str]:
    return {
        1: "厂商代码",
        2: "料號",
        3: "旧料號",
        4: "数量",
        5: "品名",
        6: "批号",
        7: "序列号",
        8: "交货日期",
        9: "生產日期",
        10: "有效日期",
        11: "订单号",
    }


def _write_label(ws, start_row: int, start_col: int, info: TemplateInfo, serial_no: int) -> None:
    sc = start_col
    sr = start_row
    c_label, c_val, c_r1, c_r2 = sc, sc + 1, sc + 2, sc + 3

    prod = _digits(info.production_date)
    year = int(prod[:4]) if len(prod) >= 4 else None
    month = prod[4:6] if len(prod) >= 6 else ""

    labels = info.field_labels or _default_field_labels()
    for off in range(1, 12):
        if off not in labels or not labels[off]:
            labels[off] = _default_field_labels()[off]

    _merge(ws, sr, c_label, sr, c_r2)
    _merge(ws, sr + 1, c_r1, sr + 4, c_r2)
    _merge(ws, sr + 5, c_val, sr + 5, c_r2)
    _merge(ws, sr + 6, c_r1, sr + 6, c_r2)
    _merge(ws, sr + 7, c_r1, sr + 8, c_r2)
    _merge(ws, sr + 9, c_r1, sr + 9, c_r2)
    _merge(ws, sr + 10, c_r1, sr + 11, c_r2)

    cell = ws.cell(sr, c_label, info.title)
    cell.font = _font(12)
    cell.alignment = _align("center")

    for off in range(1, 12):
        cell = ws.cell(sr + off, c_label, labels[off])
        cell.font = _font(12)
        cell.alignment = _align("distributed")

    # 旧料号固定留空
    values = {
        1: info.vendor,
        2: info.material_no,
        3: None,
        4: info.qty_text,
        5: info.product_name,
        6: _as_int_or_str(info.batch_no),
        7: display_serial(info, serial_no),
        8: _as_int_or_str(info.delivery_date),
        9: _as_int_or_str(info.production_date),
        10: _as_int_or_str(info.expiry_date),
        11: info.order_no,
    }
    fonts = {5: 11, 11: 10}
    aligns = {5: _align("left"), 11: _align(None)}
    for off, val in values.items():
        cell = ws.cell(sr + off, c_val, val)
        cell.font = _font(fonts.get(off, 12))
        cell.alignment = aligns.get(off, _align("center"))

    cell = ws.cell(sr + 6, c_r1, info.rohs or "RoHS")
    cell.font = _font(22)
    cell.alignment = _align("center", wrap=True)

    cell = ws.cell(sr + 7, c_r1, info.iqc or "IQC")
    cell.font = _font(22)
    cell.alignment = _align("center", wrap=True)

    if year is not None:
        cell = ws.cell(sr + 9, c_r1, year)
        cell.font = _font(20)
        cell.alignment = _align("center", wrap=True)

    if month:
        cell = ws.cell(sr + 10, c_r1, month)
        cell.font = _font(48)
        cell.alignment = _align("center", wrap=True)

    _set_border_range(ws, sr, c_label, sr + 11, c_r2)

    payload = build_qr_payload(info, serial_no)
    bio = io.BytesIO(make_qr_png_bytes(payload))
    bio.seek(0)
    xl_img = XLImage(bio)
    xl_img.width = 300
    xl_img.height = 300
    from_col = c_r1 - 1
    from_row = sr
    xl_img.anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=from_col, colOff=QR_FROM_COLOFF, row=from_row, rowOff=QR_FROM_ROWOFF),
        to=AnchorMarker(col=from_col + 1, colOff=QR_TO_COLOFF, row=from_row + 3, rowOff=QR_TO_ROWOFF),
    )
    ws.add_image(xl_img)


def _setup_sheet(ws, title: str) -> None:
    ws.sheet_view.showGridLines = False
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.scale = 78
    ws.page_margins = PageMargins(
        left=0.17, right=0.15, top=0.31, bottom=0.17, header=0.3, footer=0.17
    )
    ws.print_options.horizontalCentered = True
    safe = re.sub(r"[:\\/?*\[\]]", "_", title.strip())[:31] or "标签"
    ws.title = safe


def _ensure_row_heights(ws, start_row: int) -> None:
    for off, height in ROW_HEIGHTS.items():
        ws.row_dimensions[start_row + off].height = height


def generate_workbook(req: GenerateRequest) -> Workbook:
    serials = list(req.serials)
    if not serials:
        raise ValueError("序列号列表为空")
    if any(n < 1 for n in serials):
        raise ValueError("序列号必须全部 ≥ 1")

    info = req.template
    wb = Workbook()
    wb.remove(wb.active)

    total = len(serials)
    sheet_index = 0
    cursor = 0
    base_name = (req.sheet_name or info.material_no or info.sheet_name or "标签").strip()

    while cursor < total:
        sheet_index += 1
        n_this = min(LABELS_PER_SHEET, total - cursor)
        name = base_name if sheet_index == 1 else f"{base_name} ({sheet_index})"
        ws = wb.create_sheet(title="tmp")
        _setup_sheet(ws, name)

        for i in range(n_this):
            block = i // LABELS_PER_ROW
            col_i = i % LABELS_PER_ROW
            start_row = 2 + block * BLOCK_STRIDE
            start_col = LABEL_START_COLS[col_i]
            _ensure_row_heights(ws, start_row)
            _write_label(ws, start_row, start_col, info, serials[cursor + i])

        cursor += n_this

    return wb


def generate_to_file(req: GenerateRequest, output_path: str | Path) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = generate_workbook(req)
    wb.save(out)
    return out


def inspect_qr_in_file(path: str | Path, limit: int = 50) -> list[tuple[str, str]]:
    """读取已生成 Excel 里的二维码内容，返回 [(位置说明, 内容), ...]。"""
    wb = load_workbook(path)
    rows: list[tuple[str, str]] = []
    try:
        for ws in wb.worksheets:
            items = []
            for idx, img in enumerate(ws._images):
                try:
                    row = img.anchor._from.row
                    col = img.anchor._from.col
                    raw = img._data()
                except Exception:
                    continue
                text = _decode_png_bytes(raw)
                items.append((row, col, idx, text))
            items.sort(key=lambda x: (x[0], x[1], x[2]))
            # 去重：同一锚点可能叠了多张旧图，只留能解出的第一张
            seen = set()
            for row, col, idx, text in items:
                key = (row, col)
                if key in seen:
                    continue
                if not text:
                    continue
                seen.add(key)
                serial = text.split(";")[-1] if ";" in text else ""
                rows.append((f"{ws.title} 行列({row+1},{col+1}) 末尾序列={serial}", text))
                if len(rows) >= limit:
                    return rows
    finally:
        wb.close()
    return rows


def planned_qr_list(info: TemplateInfo, serials: list[int]) -> list[tuple[str, str]]:
    """生成前预览：显示序列号 ↔ 二维码全文。"""
    return [(display_serial(info, n), build_qr_payload(info, n)) for n in serials]
