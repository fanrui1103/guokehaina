"""解析送货单明细 Excel：只统计「已发货」，按物料编码分组。"""

from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

# 只处理这一状态；其它如「已结案」「关闭」一律跳过
SHIPPED_STATUS = "已发货"


@dataclass
class MaterialGroup:
    part_no: str
    name: str
    qty: int
    po_no: str
    row_count: int


def _norm_header(v) -> str:
    return str(v).strip() if v is not None else ""


def _to_qty(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_int_qty(total: float) -> int:
    if abs(total - round(total)) < 1e-9:
        return int(round(total))
    return int(total)


def parse_delivery_excel(path: str | Path) -> list[MaterialGroup]:
    """
    读取送货单明细（仅「数据状态 = 已发货」的行）。
    - 物料编码：分组键
    - 发货数量：同组求和 → 订单数量/交货数量
    - 采购订单号：取该组第一行
    """
    path = Path(path)
    # 注意：部分企微导出的 xlsx 在 read_only 下表头会缺列，必须普通打开
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_norm_header(h) for h in rows[0]]
        rows = rows[1:]

        def col(*names: str) -> int:
            for n in names:
                if n in headers:
                    return headers.index(n)
            raise KeyError("缺少列：" + " / ".join(names))

        i_mat = col("物料编码", "物料编号")
        i_qty = col("发货数量")
        i_po = col("采购订单号")
        i_status = col("数据状态")
        i_name = headers.index("物料名称") if "物料名称" in headers else None

        groups: OrderedDict[str, dict] = OrderedDict()
        for r in rows:
            if not r or i_mat >= len(r) or r[i_mat] is None:
                continue
            status = _norm_header(r[i_status] if i_status < len(r) else "")
            if status != SHIPPED_STATUS:
                continue

            part_no = str(r[i_mat]).strip()
            if not part_no:
                continue
            qty = _to_qty(r[i_qty] if i_qty < len(r) else None)
            po = r[i_po] if i_po < len(r) else None
            po_s = str(po).strip() if po is not None and str(po).strip() else ""
            name = ""
            if i_name is not None and i_name < len(r) and r[i_name] is not None:
                name = str(r[i_name]).strip()

            if part_no not in groups:
                groups[part_no] = {
                    "name": name,
                    "qty": 0.0,
                    "po_no": po_s,
                    "rows": 0,
                }
            groups[part_no]["qty"] += qty
            groups[part_no]["rows"] += 1
            if not groups[part_no]["name"] and name:
                groups[part_no]["name"] = name

        result: list[MaterialGroup] = []
        for part_no, g in groups.items():
            result.append(
                MaterialGroup(
                    part_no=part_no,
                    name=g["name"] or part_no,
                    qty=_to_int_qty(g["qty"]),
                    po_no=g["po_no"],
                    row_count=g["rows"],
                )
            )
        return result
    finally:
        wb.close()
