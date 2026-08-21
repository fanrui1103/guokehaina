"""把识别结果写入雅达物料清单模板。"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
COL = {
    "PONo": "A",
    "POLine": "B",
    "InvoiceDN": "C",
    "AirWayBill": "D",
    "BillOfLading": "E",
    "TruckNo": "F",
    "AdditionalText": "G",
    "StoragePosition": "H",
    "DeliveryType": "I",
    "Operator": "J",
    "ManufactureDate": "K",
    "DateCode": "L",
    "LotNo": "M",
    "COO": "N",
    "UnitQty": "O",
    "Unit": "P",
    "NoOfPackage": "Q",
    "supplier": "R",
    "supplier_code": "S",
    "Manufacturer": "T",
    "ManufacturerPN": "U",
    "CLID": "V",
    "ExpDate": "W",
    "ArtesynPN": "X",
    "OldAstecPN": "Y",
    "MaterialGroup": "Z",
}

DATE_FIELDS = {"ManufactureDate", "DateCode", "ExpDate"}
YELLOW = PatternFill("solid", fgColor="FFFF00")
FONT = Font(name="Times New Roman", size=12)
ALIGN = Alignment(vertical="center")


def _as_excel_value(key: str, value):
    if value is None or value == "":
        return None
    if key in DATE_FIELDS:
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            text = value.strip().replace("-", "/").replace(".", "/")
            try:
                return datetime.strptime(text, "%Y/%m/%d")
            except ValueError:
                pass
            parts = text.split("/")
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
        return value
    return value


def write_excel(template_path: str | Path, output_path: str | Path, rows: list[dict]) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if template_path.resolve() != output_path.resolve():
        shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    if ws["A1"].value != "订单号":
        ws["A1"] = "订单号"

    start_row = 3
    for offset, row in enumerate(rows):
        r = start_row + offset
        for key, col in COL.items():
            cell = ws[f"{col}{r}"]
            value = _as_excel_value(key, row.get(key))
            cell.value = value
            cell.font = FONT
            cell.alignment = ALIGN
            if key in DATE_FIELDS:
                cell.number_format = "yyyy/m/d"
            if key == "CLID" and value:
                cell.fill = YELLOW

    # 清掉模板里多出来的空格式行以外的旧内容（若复制的是已填过的表）
    last = start_row + len(rows) - 1
    if ws.max_row > last:
        for r in range(last + 1, ws.max_row + 1):
            for c in range(1, 27):
                ws.cell(r, c).value = None

    wb.save(output_path)
    wb.close()
    return output_path


def default_output_name() -> str:
    return datetime.now().strftime("物料清单_%Y%m%d_%H%M.xlsx")
